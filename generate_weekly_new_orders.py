#!/usr/bin/env python3
"""Generate 「6.周新客订单表」 rows from ERP API + local Excel."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import urllib.parse
import urllib.request
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

try:
    import yaml
except ImportError:
    yaml = None  # type: ignore


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def load_config(path: Path) -> dict:
    if yaml is None:
        raise SystemExit("PyYAML required: pip install pyyaml")
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_brands(path: Path) -> dict:
    with path.open(encoding="utf-8") as f:
        return yaml.safe_load(f)


# ---------------------------------------------------------------------------
# Dates
# ---------------------------------------------------------------------------

def parse_api_date(s: str) -> date:
    return datetime.strptime(s[:10], "%Y-%m-%d").date()


def to_excel_serial(d: date) -> int:
    return (d - date(1899, 12, 30)).days


def serial_to_date(serial: int | float | str) -> date | None:
    try:
        return date(1899, 12, 30) + timedelta(days=int(float(serial)))
    except (TypeError, ValueError):
        return None


def period_label(begin: date, end: date) -> str:
    return f"{begin.month:02d}.{begin.day:02d} - {end.month:02d}.{end.day:02d}"


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

class YouroApi:
    def __init__(self, base_url: str, jsessionid: str):
        self.base_url = base_url.rstrip("/")
        self.cookie = f"JSESSIONID={jsessionid}"

    def _post(self, path: str, data: dict[str, str]) -> dict:
        body = urllib.parse.urlencode(data).encode()
        req = urllib.request.Request(
            f"{self.base_url}{path}",
            data=body,
            headers={
                "Content-Type": "application/x-www-form-urlencoded",
                "X-Requested-With": "XMLHttpRequest",
                "Cookie": self.cookie,
                "User-Agent": "weekly-new-orders/1.0",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=60) as resp:
            payload = json.loads(resp.read().decode())
        if payload.get("code") != 0:
            raise RuntimeError(f"API error {path}: {payload}")
        return payload

    def list_sales_orders(self, begin: str, end: str, page_size: int = 100) -> list[dict]:
        data = {
            "pageSize": str(page_size),
            "pageNum": "1",
            "orderByColumn": "createTime",
            "isAsc": "desc",
            "orderNo": "",
            "customerId": "",
            "customerName": "",
            "country": "",
            "trackerNo": "",
            "productName": "",
            "shipmentStatus": "",
            "params[beginOrderDate]": begin,
            "params[endOrderDate]": end,
            "createBy": "",
            "purchaser": "",
            "purchaseStatus": "",
            "payMethod": "",
            "currency": "",
            "receiptCheck": "",
            "freightCheck": "",
            "warehouse": "",
            "company": "",
            "isUrgent": "",
            "firstOrder": "Y",
            "statusArrays": "",
        }
        return self._post("/comp/sales-order/list", data).get("rows", [])

    def get_purchaser_order(self, order_no: str) -> dict | None:
        data = {
            "pageSize": "10",
            "pageNum": "1",
            "orderByColumn": "createTime",
            "isAsc": "desc",
            "orderNo": order_no,
            "customerId": "",
            "customerName": "",
            "country": "",
            "trackerNo": "",
            "productName": "",
            "payMethod": "",
            "shipmentStatus": "",
            "params[beginOrderDate]": "2020-01-01",
            "params[endOrderDate]": "2030-12-31",
            "createBy": "",
            "purchaser": "",
            "purchaseStatus": "",
            "warehouse": "",
            "isUrgent": "",
            "firstOrder": "",
            "statusArrays": "",
        }
        rows = self._post("/comp/purchaser-order/list", data).get("rows", [])
        return rows[0] if rows else None


# ---------------------------------------------------------------------------
# Excel (zip/xml — avoids openpyxl compatibility issues on large files)
# ---------------------------------------------------------------------------

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"


def _col_index(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch.upper()) - ord("A") + 1)
    return n - 1


def read_xlsx_rows(xlsx_path: Path, sheet_name: str | None = None) -> dict[str, list[list[Any]]]:
    """Return {sheet_name: [row_values,...]} row index 0 = Excel row 1."""
    with zipfile.ZipFile(xlsx_path) as zf:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall(".//m:si", NS):
                shared.append("".join(t.text or "" for t in si.findall(".//m:t", NS)))

        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = {
            rel.get("Id"): rel.get("Target")
            for rel in ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        }
        sheet_paths: dict[str, str] = {}
        for sh in wb.findall(".//m:sheets/m:sheet", NS):
            name = sh.get("name")
            rid = sh.get(REL_NS + "id")
            target = rels.get(rid, "")
            sheet_paths[name] = "xl/" + target.lstrip("/")

        result: dict[str, list[list[Any]]] = {}
        names = [sheet_name] if sheet_name else list(sheet_paths.keys())
        for name in names:
            if name not in sheet_paths:
                continue
            root = ET.fromstring(zf.read(sheet_paths[name]))
            max_col = 0
            sparse: dict[int, dict[int, Any]] = {}
            for row_el in root.findall(".//m:sheetData/m:row", NS):
                r_idx = int(row_el.get("r")) - 1
                sparse[r_idx] = {}
                for c in row_el.findall("m:c", NS):
                    ref = c.get("r", "")
                    m = re.match(r"([A-Z]+)", ref)
                    if not m:
                        continue
                    ci = _col_index(m.group(1))
                    max_col = max(max_col, ci)
                    v_el = c.find("m:v", NS)
                    if v_el is None or v_el.text is None:
                        continue
                    val: Any = v_el.text
                    if c.get("t") == "s":
                        val = shared[int(val)]
                    elif re.fullmatch(r"-?\d+(\.\d+)?", val):
                        val = float(val) if "." in val else int(val)
                    sparse[r_idx][ci] = val

            if not sparse:
                result[name] = []
                continue
            last_row = max(sparse.keys())
            rows_out: list[list[Any]] = []
            for ri in range(last_row + 1):
                row_map = sparse.get(ri, {})
                width = max(max_col + 1, max(row_map.keys(), default=-1) + 1)
                rows_out.append([row_map.get(ci, "") for ci in range(width)])
            result[name] = rows_out
        return result


def _norm(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


def _find_sheet(rows_by_sheet: dict[str, list[list[Any]]], *keywords: str) -> tuple[str, list[list[Any]]]:
    for name, rows in rows_by_sheet.items():
        if all(k in name for k in keywords):
            return name, rows
    for name, rows in rows_by_sheet.items():
        if any(k in name for k in keywords):
            return name, rows
    raise KeyError(f"No sheet matching {keywords}")


@dataclass
class A02OrderRow:
    order_date: date | None = None
    sales: str = ""
    customer: str = ""
    shop: str = ""
    country: str = ""
    qty: int | float | None = None
    order_amount_usd: float | None = None
    payment_usd: float | None = None
    payment_rmb: float | None = None
    product_rmb: float | None = None
    pay_channel: str = ""
    purchase_rmb: float | None = None
    freight_rmb: float | None = None
    other_freight_rmb: float | None = None
    logistics: str = ""
    initial_freight: float | None = None
    gross_profit: float | None = None
    gross_margin: float | None = None


@dataclass
class TrafficRow:
    sales: str = ""
    customer: str = ""
    country: str = ""
    add_date: date | None = None
    source: str = ""
    traffic_type: str = ""
    category: str = ""
    level: str = ""
    shop: str = ""
    file: str = ""
    source_workbook: str = ""
    source_sheet: str = ""


def is_a05_traffic(t: TrafficRow) -> bool:
    if t.source_workbook.startswith("A05"):
        return True
    return t.file.startswith("A05/")


def format_traffic_source(t: TrafficRow) -> str:
    if t.source_workbook and t.source_sheet:
        return f"{t.source_workbook} › {t.source_sheet}"
    if t.source_workbook:
        return t.source_workbook
    return t.file


def summarize_traffic_sources(rows: list[TrafficRow]) -> str:
    if not rows:
        return "（无数据）"
    from collections import Counter
    counts = Counter(format_traffic_source(t) for t in rows)
    return "；".join(f"「{src}」{n}条" for src, n in counts.most_common())


def _float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_a02_orders(a02_path: Path) -> dict[tuple, A02OrderRow]:
    """By (date, sales, customer) from 屿路 - 订单信息表."""
    all_sheets = read_xlsx_rows(a02_path)
    _, order_rows = _find_sheet(all_sheets, "订单信息")

    by_match: dict[tuple, A02OrderRow] = {}

    for row in order_rows[3:]:  # skip title/header rows
        if len(row) < 4 or not row[1]:
            continue
        d = serial_to_date(row[1])
        sales = str(row[2] or "")
        customer = str(row[3] or "")
        rec = A02OrderRow(
            order_date=d,
            sales=sales,
            customer=customer,
            shop=str(row[5] if len(row) > 5 else ""),
            country=str(row[6] if len(row) > 6 else ""),
            qty=row[7] if len(row) > 7 else None,
            order_amount_usd=_float(row[8] if len(row) > 8 else None),
            payment_usd=_float(row[10] if len(row) > 10 else None),
            payment_rmb=_float(row[11] if len(row) > 11 else None),
            product_rmb=_float(row[13] if len(row) > 13 else None),
            pay_channel=str(row[14] if len(row) > 14 else ""),
            purchase_rmb=_float(row[15] if len(row) > 15 else None),
            freight_rmb=_float(row[16] if len(row) > 16 else None),
            other_freight_rmb=_float(row[17] if len(row) > 17 else None),
            logistics=str(row[18] if len(row) > 18 else ""),
            initial_freight=_float(row[19] if len(row) > 19 else None),
            gross_profit=_float(row[20] if len(row) > 20 else None),
            gross_margin=_float(row[21] if len(row) > 21 else None),
        )
        if d and sales and customer:
            by_match[(d.isoformat(), _norm(sales), _norm(customer))] = rec

    return by_match


MULTI_SHOP_A060X_PREFIXES = frozenset({"A0604", "A0607"})


def build_a05_shop_index(traffic_rows: list[TrafficRow]) -> dict[tuple[str, str, str], str]:
    """(sales, customer, add_date) -> 屿路/镕川 from A05 authoritative sheets."""
    index: dict[tuple[str, str, str], str] = {}
    for t in traffic_rows:
        if not is_a05_traffic(t) or not t.add_date or not t.shop:
            continue
        key = (_norm(t.sales), _norm(t.customer), t.add_date.isoformat())
        index[key] = t.shop
    return index


def infer_a060x_shop(
    workbook_prefix: str,
    sales: str,
    customer: str,
    add_date: date | None,
    traffic_type: str,
    a05_index: dict[tuple[str, str, str], str],
    default_shop: str,
) -> str:
    """A0604 Grace / A0607 Lily 单表含 Youro+RonChamp，按 A05 对齐 + 类型兜底。"""
    if add_date:
        key = (_norm(sales), _norm(customer), add_date.isoformat())
        if key in a05_index:
            return a05_index[key]

    tt = str(traffic_type or "").strip().upper()
    if workbook_prefix == "A0604" or _norm(sales) == _norm("Grace"):
        if tt == "R-TM" or "R-TM" in tt:
            return "镕川"
        if tt == "TM":
            return "屿路"
    if workbook_prefix == "A0607" or _norm(sales) == _norm("Lily"):
        if tt in ("询盘", "RFQ"):
            return "镕川"
        if tt == "TM":
            return "镕川"
    return default_shop


def load_traffic_rows(data_dir: Path, a05_path: Path, globs: list[str]) -> list[TrafficRow]:
    rows_out: list[TrafficRow] = []

    def parse_traffic_sheet(
        rows: list[list[Any]],
        workbook: str,
        sheet: str,
        default_shop: str = "",
        *,
        a05_index: dict[tuple[str, str, str], str] | None = None,
        workbook_prefix: str = "",
    ) -> None:
        if len(rows) < 3:
            return
        header = [str(x) for x in rows[1]]
        if "客户姓名" not in "".join(header) and "客户" not in "".join(header):
            return
        multi_shop = workbook_prefix in MULTI_SHOP_A060X_PREFIXES
        for row in rows[2:]:
            if len(row) < 6:
                continue
            customer = str(row[2] if len(row) > 2 else "").strip()
            if not customer:
                continue
            add_d = serial_to_date(row[5] if len(row) > 5 else None)
            sales = str(row[1] if len(row) > 1 else "")
            traffic_type = str(row[8] if len(row) > 8 else "")
            if multi_shop and a05_index is not None:
                shop = infer_a060x_shop(
                    workbook_prefix, sales, customer, add_d, traffic_type, a05_index, default_shop
                )
            else:
                shop = default_shop
            rows_out.append(
                TrafficRow(
                    sales=sales,
                    customer=customer,
                    country=str(row[4] if len(row) > 4 else ""),
                    add_date=add_d,
                    source=str(row[7] if len(row) > 7 else ""),
                    traffic_type=traffic_type,
                    category=str(row[9] if len(row) > 9 else ""),
                    level=str(row[6] if len(row) > 6 else ""),
                    shop=shop,
                    file=f"{workbook} › {sheet}",
                    source_workbook=workbook,
                    source_sheet=sheet,
                )
            )

    a05_workbook = a05_path.name
    a05_sheets = read_xlsx_rows(a05_path)
    for name, rows in a05_sheets.items():
        if "新流量" in name:
            shop = "屿路" if "屿路" in name else "镕川" if "镕川" in name else ""
            parse_traffic_sheet(rows, a05_workbook, name, shop)

    a05_index = build_a05_shop_index(rows_out)

    shop_from_file = {
        "A0602": "屿路", "A0603": "屿路", "A0605": "屿路",
        "A0601": "屿路", "A0606": "镕川",
        # A0604 Grace / A0607 Lily：表内混 Youro+RonChamp，用 infer_a060x_shop
        "A0604": "屿路", "A0607": "镕川",
    }
    for pattern in globs:
        for path in sorted(data_dir.glob(pattern)):
            if path.name.startswith("~$"):
                continue
            prefix = path.name.split("-")[0]
            default_shop = shop_from_file.get(prefix, "")
            try:
                for name, rows in read_xlsx_rows(path).items():
                    if "新流量" in name or name.strip() in ("新流量表",):
                        parse_traffic_sheet(
                            rows,
                            path.name,
                            name,
                            default_shop,
                            a05_index=a05_index if prefix in MULTI_SHOP_A060X_PREFIXES else None,
                            workbook_prefix=prefix,
                        )
            except Exception as exc:
                print(f"  warn: skip {path.name}: {exc}", file=sys.stderr)

    return rows_out


def find_traffic(
    traffic_rows: list[TrafficRow],
    sales: str,
    customer: str,
    order_date: date,
    store: str | None = None,
) -> TrafficRow | None:
    ns, nc = _norm(sales), _norm(customer)
    candidates = [
        t
        for t in traffic_rows
        if _norm(t.customer) == nc and (_norm(t.sales) == ns or not t.sales)
    ]
    if store:
        shop = STORE_TO_SHOP.get(store, "")
        store_matched = [t for t in candidates if not t.shop or t.shop == shop or shop_to_store(t.shop) == store]
        if store_matched:
            candidates = store_matched
    if not candidates:
        candidates = [t for t in traffic_rows if _norm(t.customer) == nc]
    if not candidates:
        return None
    before = [t for t in candidates if t.add_date and t.add_date <= order_date]
    pool = before or candidates
    pool.sort(key=lambda t: t.add_date or date.min, reverse=True)
    pool.sort(key=lambda t: (0 if is_a05_traffic(t) else 1))
    return pool[0]


# ---------------------------------------------------------------------------
# Brand resolver
# ---------------------------------------------------------------------------

@dataclass
class BrandResult:
    brand: str
    confidence: str
    source: str


def resolve_brand(product_name: str, category: str, rules: dict) -> BrandResult:
    text = product_name or ""
    upper = text.upper()
    hits: list[tuple[str, str]] = []

    for prefix, brand in (rules.get("prefixes") or {}).items():
        if prefix.upper() in upper:
            hits.append((brand, "prefix"))

    for kw, brand in (rules.get("keywords") or {}).items():
        if kw.upper() in upper or kw in text:
            hits.append((brand, "keyword"))

    cat_brand = None
    for kw, brand in (rules.get("category") or {}).items():
        if kw in (category or ""):
            cat_brand = brand
            break

    hits = [(b, s) for b, s in hits if b != "待确认"]
    if hits:
        brands = {h[0] for h in hits}
        if len(brands) == 1:
            b = next(iter(brands))
            conf = "高" if cat_brand == b else "中"
            return BrandResult(b, conf, hits[0][1])
        return BrandResult("待确认", "低", "conflict")

    if cat_brand:
        return BrandResult(cat_brand, "中", "category")

    return BrandResult("待确认", "待确认", "fallback")


def normalize_sales(name: str, mapping: dict) -> str:
    key = _norm(name)
    if key in mapping:
        return mapping[key]
    return name.strip().title() if name else ""


def pay_channel(api_method: Any, a02: A02OrderRow | None, pay_map: dict) -> str:
    if a02 and a02.pay_channel:
        return a02.pay_channel
    return pay_map.get(str(api_method), "")


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

STORE_YOURO = "Youro"
STORE_RONCHAMP = "RonChamp"
SHOP_TO_STORE = {"屿路": STORE_YOURO, "镕川": STORE_RONCHAMP}
STORE_TO_SHOP = {STORE_YOURO: "屿路", STORE_RONCHAMP: "镕川"}

YOURO_SALES_ORDER = ["Ennerson", "Luck", "Cindy", "Grace", "David", "Lily"]
RONCHAMP_SALES_ORDER = ["Lily", "Sally", "Grace"]


def shop_to_store(shop: str) -> str:
    return SHOP_TO_STORE.get(str(shop or "").strip(), STORE_YOURO)


def store_to_label(store: str) -> str:
    return "RonChamp" if store == STORE_RONCHAMP else "Youro"


def is_l1plus(level: Any) -> bool:
    s = str(level or "").strip().upper().replace(" ", "")
    return s not in ("", "L0") and s != "L1-"


OTHER_CHANNEL_KEYWORDS = ("转介绍", "微信", "公海", "客户介绍", "介绍", "老客户")


def _text_has_other_keyword(text: str) -> bool:
    return any(kw in str(text or "") for kw in OTHER_CHANNEL_KEYWORDS)


def is_explicit_other_channel(traffic: "TrafficRow | None", sale: dict | None = None) -> bool:
    if traffic:
        combined = f"{traffic.source or ''}{traffic.traffic_type or ''}"
        if _text_has_other_keyword(combined):
            return True
    if sale:
        for field in ("salesRemark", "remark", "receiptRemark"):
            if _text_has_other_keyword(str(sale.get(field) or "")):
                return True
    return False


def classify_channel(traffic: "TrafficRow | None", sale: dict | None = None) -> str:
    """tm / rfq / other（明确非 TM 来源）/ unclassified（无流量且非明确其他）."""
    if traffic:
        combined = f"{traffic.source or ''}{traffic.traffic_type or ''}"
        if _text_has_other_keyword(combined):
            return "other"
        t = str(traffic.traffic_type or "").upper()
        s = str(traffic.source or "").upper()
        if "RFQ" in t or "RFQ" in s:
            return "rfq"
        if t in ("TM", "询盘", "R-TM") or "TM" in t or "询盘" in t:
            return "tm"
        if traffic.source or traffic.traffic_type:
            return "tm"
    if is_explicit_other_channel(traffic, sale):
        return "other"
    return "unclassified"


def resolve_store(a02: A02OrderRow | None, sale: dict) -> str:
    if a02 and str(a02.shop or "").strip():
        return shop_to_store(a02.shop)
    company = sale.get("company")
    if company in (1, "1"):
        return STORE_RONCHAMP
    return STORE_YOURO


@dataclass
class OrderMetrics:
    order_no: str
    store: str
    sales: str
    channel: str
    payment_rmb: float
    gross: float | None
    margin: float | None
    customer: str = ""
    order_date: str = ""
    traffic_file: str = ""
    traffic_source: str = ""


def compute_order_metrics(
    sale: dict,
    purchaser: dict | None,
    a02: A02OrderRow | None,
    traffic: TrafficRow | None,
    brands: dict,
    cfg: dict,
) -> OrderMetrics:
    row, _, _ = build_row(sale, purchaser, a02, traffic, brands, cfg)
    sales = str(row[2])
    payment_rmb = float(row[8]) if row[8] != "" else 0.0
    gross = float(row[16]) if row[16] != "" else None
    margin = float(row[17]) if row[17] != "" else None
    store = resolve_store(a02, sale)
    channel = classify_channel(traffic, sale)
    return OrderMetrics(
        order_no=str(sale.get("orderNo", "")),
        store=store,
        sales=sales,
        channel=channel,
        payment_rmb=payment_rmb,
        gross=gross,
        margin=margin,
        customer=str(sale.get("customerName", "")),
        order_date=str(sale.get("orderDate", ""))[:10],
        traffic_file=format_traffic_source(traffic) if traffic else "",
        traffic_source=f"{traffic.source or ''}/{traffic.traffic_type or ''}" if traffic else "",
    )


CONVERSION_HEADERS = [
    "所属店铺", "业务员", "总成交数",
    "TM_6月新客数量", "TM_L1+新客数量", "TM_L1+占比",
    "TM_本月成交新客数量", "TM_新客转化率", "TM_新客成交金额", "TM_新客毛利润", "TM_新客毛利率",
    "RFQ_本月成交新客数量", "RFQ_成交金额", "RFQ_新客毛利润", "RFQ_新客毛利率",
    "其他_本月成交新客数量", "其他_成交金额", "其他_新客毛利润", "其他_新客毛利率",
    "备注",
]

WEEKLY_HEADERS = [
    "周期", "订单日期", "销售人员", "客户姓名", "所属国家", "型号数量",
    "订单总金额（美金）", "实际到帐金额（美金）", "实际到帐金额（人民币）",
    "产品总金额（人民币）", "收款渠道", "采购金额（人民币）", "运费（人民币）",
    "其它运输运费", "物流方式", "初始运费", "毛利润（人民币）", "毛利率",
    "流量添加日期", "流量来源", "咨询品类", "成交产品",
]


@dataclass
class ChannelReview:
    order_no: str
    store: str
    sales: str
    customer: str
    order_date: str
    payment_rmb: float
    gross: float | None
    traffic_file: str
    traffic_source: str
    note: str


@dataclass
class PurchaseReview:
    order_no: str
    purchase_api: float | None
    purchase_a02_order: float | None
    purchase_diff: str
    purchase_alert: str


def compare_purchase(order_no: str, api: float | None, a02: float | None) -> PurchaseReview:
    vals = {"API": api, "A02订单": a02}
    present = {k: v for k, v in vals.items() if v is not None}
    alerts: list[str] = []
    if len(present) < 2:
        diff_flag = "部分缺失" if present else "缺失"
        if len(present) == 1:
            alerts.append(f"仅有一路数据: {list(present.items())[0]}")
    else:
        amounts = list(present.values())
        if max(amounts) - min(amounts) > 0.01:
            diff_flag = "是"
            alerts.append(", ".join(f"{k}({v})" for k, v in present.items()))
        else:
            diff_flag = "否"
    return PurchaseReview(
        order_no=order_no,
        purchase_api=api,
        purchase_a02_order=a02,
        purchase_diff=diff_flag,
        purchase_alert="; ".join(alerts) if alerts else "",
    )


@dataclass
class BrandReview:
    order_no: str
    brand_suggested: str
    brand_confidence: str
    brand_source: str
    productName_raw: str
    category_raw: str
    brand_final: str = ""


def build_row(
    sale: dict,
    purchaser: dict | None,
    a02: A02OrderRow | None,
    traffic: TrafficRow | None,
    brands: dict,
    cfg: dict,
) -> tuple[list[Any], PurchaseReview, BrandReview]:
    order_no = sale["orderNo"]
    order_date = parse_api_date(sale["orderDate"])
    sales = normalize_sales(sale.get("createBy", ""), cfg.get("sales_name_map", {}))
    customer = str(sale.get("customerName", "")).strip()
    country = sale.get("country", "")
    qty = sale.get("qty", "")
    order_amount = float(sale.get("orderAmount") or 0)
    payment_usd = float(sale.get("paymentReceived") or 0)
    product_usd = float(sale.get("productAmount") or 0)
    rate = float(sale.get("exchangeRate") or 6.7)

    payment_rmb = round(payment_usd * rate, 2)
    product_rmb = round(product_usd * rate, 2)

    purchase_api = _float(purchaser.get("purchaseAmount") if purchaser else None)
    purchase_a02 = a02.purchase_rmb if a02 else None
    purchase_review = compare_purchase(order_no, purchase_api, purchase_a02)

    if purchase_api is not None:
        purchase_rmb = purchase_api
    elif purchase_a02 is not None:
        purchase_rmb = purchase_a02
    else:
        purchase_rmb = None

    freight = a02.freight_rmb if a02 else None
    other_freight = a02.other_freight_rmb if a02 else None
    logistics = (a02.logistics if a02 and a02.logistics else "") or str(sale.get("salesRemark") or "").strip()
    # P 列仅写 A02 初始运费；无 A02 时不从 API 回填（与历史周表习惯一致）
    initial_freight = a02.initial_freight if a02 else None

    channel = pay_channel(sale.get("payMethod"), a02, cfg.get("pay_method_map", {}))

    # 毛利润：默认 Q=J−L；仅当 A02 有运费分项时才扩展扣减
    m = freight or 0
    n = other_freight or 0
    o = 0  # 物流方式列为文本，不参与数值扣减
    p = initial_freight or 0
    has_a02_freight = any(v not in (None, "", 0) for v in (freight, other_freight, initial_freight))

    if purchase_rmb is not None:
        if has_a02_freight:
            gross = round(product_rmb - purchase_rmb - m - n - p, 2)
        else:
            gross = round(product_rmb - purchase_rmb, 2)
        margin = round(gross / product_rmb, 4) if product_rmb else None
    else:
        gross = None
        margin = None

    if a02 and a02.gross_profit is not None and gross is not None:
        if abs(a02.gross_profit - gross) / max(abs(gross), 1) < 0.01:
            gross = a02.gross_profit
            margin = a02.gross_margin

    traffic_type = traffic.traffic_type if traffic else ""
    if not traffic_type and traffic:
        traffic_type = traffic.source
    # 无流量匹配时不默认填「转介绍」；仅 classify_channel=other 时在转化逻辑中处理
    category = traffic.category if traffic else ""
    add_serial = to_excel_serial(traffic.add_date) if traffic and traffic.add_date else ""

    product_name = str(sale.get("productName") or "")
    if purchaser and purchaser.get("productName"):
        product_name = product_name or str(purchaser["productName"])
    brand_res = resolve_brand(product_name, category, brands)
    deal_product = f"成交产品：{brand_res.brand}"

    begin = parse_api_date(cfg["week"]["begin_date"])
    end = parse_api_date(cfg["week"]["end_date"])
    period = cfg["week"].get("period_label") or period_label(begin, end)

    row = [
        period,
        to_excel_serial(order_date),
        sales,
        customer,
        country,
        qty,
        order_amount,
        payment_usd,
        payment_rmb,
        product_rmb,
        channel,
        purchase_rmb if purchase_rmb is not None else "",
        freight if freight is not None else "",
        other_freight if other_freight is not None else "",
        logistics,
        initial_freight if initial_freight is not None else "",
        gross if gross is not None else "",
        margin if margin is not None else "",
        add_serial,
        traffic_type,
        category,
        deal_product,
    ]

    brand_review = BrandReview(
        order_no=order_no,
        brand_suggested=brand_res.brand,
        brand_confidence=brand_res.confidence,
        brand_source=brand_res.source,
        productName_raw=product_name[:500],
        category_raw=category,
    )
    return row, purchase_review, brand_review


def write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)


def write_conversion_csv(path: Path, title: str, rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow([conversion_title_full(title)] + [""] * (len(CONVERSION_HEADERS) - 1))
        w.writerow(CONVERSION_HEADERS)
        w.writerows(rows)


def write_review_csv(path: Path, headers: list[str], objects: list[Any]) -> None:
    rows = [[getattr(o, h) if hasattr(o, h) else o.get(h) for h in headers] for o in objects]
    write_csv(path, headers, rows)


def patch_weekly_xlsx(
    xlsx_path: Path,
    new_rows: list[list[Any]],
    period: str,
    sheet_prefix: str | None = None,
) -> None:
    """Append/replace rows for target period in weekly new-order sheet."""
    try:
        import openpyxl
    except ImportError:
        print("  skip xlsx writeback: pip install openpyxl", file=sys.stderr)
        return

    wb = openpyxl.load_workbook(xlsx_path)
    if sheet_prefix:
        sheet_name = next(
            (
                n
                for n in wb.sheetnames
                if n.startswith(sheet_prefix) or (sheet_prefix in n and "新客订单" in n)
            ),
            None,
        )
    else:
        sheet_name = next((n for n in wb.sheetnames if "新客订单" in n), None)
    if not sheet_name:
        print(f"  skip xlsx writeback: sheet not found (prefix={sheet_prefix})", file=sys.stderr)
        return
    ws = wb[sheet_name]

    to_delete = []
    for r in range(4, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).strip() == period:
            to_delete.append(r)
    for r in reversed(to_delete):
        ws.delete_rows(r, 1)

    start = ws.max_row + 1
    for i, row in enumerate(new_rows):
        for j, val in enumerate(row, 1):
            ws.cell(start + i, j, val)
    wb.save(xlsx_path)
    print(f"  wrote {len(new_rows)} rows to {xlsx_path.name} ({sheet_name})")


def _sales_sort_key(sales: str, order_list: list[str]) -> tuple[int, str]:
    try:
        return (order_list.index(sales), sales)
    except ValueError:
        return (len(order_list), sales)


def _traffic_in_range(t: TrafficRow, store: str, sales: str, month_begin: date, month_end: date) -> bool:
    shop = STORE_TO_SHOP.get(store, "")
    if _norm(t.sales) != _norm(sales):
        return False
    if t.shop and shop_to_store(t.shop) != store and t.shop != shop:
        return False
    if not t.add_date or not (month_begin <= t.add_date <= month_end):
        return False
    return True


def _traffic_row_key(t: TrafficRow) -> tuple[str, str]:
    return (_norm(t.customer), t.add_date.isoformat() if t.add_date else "")


@dataclass
class TrafficCrossSummary:
    store: str
    sales: str
    period: str
    a05_source: str
    a060x_source: str
    a05_count: int
    a060x_count: int
    overlap: int
    a05_only: int
    a060x_only: int
    field_mismatch: int
    diff: str
    alert: str


@dataclass
class TrafficCrossDetail:
    store: str
    sales: str
    customer: str
    status: str
    add_date_a05: str
    add_date_a060x: str
    level_a05: str
    level_a060x: str
    a05_workbook: str
    a05_sheet: str
    a060x_workbook: str
    a060x_sheet: str
    alert: str


def compare_traffic_cross(
    traffic_rows: list[TrafficRow],
    month_begin: date,
    month_end: date,
    sales_map: dict,
) -> tuple[list[TrafficCrossSummary], list[TrafficCrossDetail]]:
    """Cross-check A05 shop traffic vs A060x salesperson traffic for the same period."""
    period = f"{month_begin.month}.{month_begin.day} - {month_end.month}.{month_end.day}"
    summaries: list[TrafficCrossSummary] = []
    details: list[TrafficCrossDetail] = []

    pairs: set[tuple[str, str]] = set()
    for t in traffic_rows:
        if not t.add_date or not (month_begin <= t.add_date <= month_end):
            continue
        if not t.shop:
            continue
        store = shop_to_store(t.shop)
        sales = normalize_sales(t.sales, sales_map)
        if sales:
            pairs.add((store, sales))

    for store, sales in sorted(pairs):
        a05 = [
            t for t in traffic_rows
            if is_a05_traffic(t) and _traffic_in_range(t, store, sales, month_begin, month_end)
        ]
        a060x = [
            t for t in traffic_rows
            if not is_a05_traffic(t) and _traffic_in_range(t, store, sales, month_begin, month_end)
        ]
        if not a05 and not a060x:
            continue

        a05_source_desc = summarize_traffic_sources(a05)
        a060x_source_desc = summarize_traffic_sources(a060x)

        a05_by_key = {_traffic_row_key(t): t for t in a05}
        a060x_by_key = {_traffic_row_key(t): t for t in a060x}
        overlap_keys = set(a05_by_key) & set(a060x_by_key)
        a05_only_keys = set(a05_by_key) - overlap_keys
        a060x_only_keys = set(a060x_by_key) - overlap_keys

        field_mismatch = 0
        alerts: list[str] = []

        for key in sorted(overlap_keys):
            ta, tb = a05_by_key[key], a060x_by_key[key]
            la, lb = str(ta.level or "").strip(), str(tb.level or "").strip()
            if la != lb:
                field_mismatch += 1
                details.append(
                    TrafficCrossDetail(
                        store=store_to_label(store),
                        sales=sales,
                        customer=ta.customer,
                        status="字段不一致",
                        add_date_a05=ta.add_date.isoformat() if ta.add_date else "",
                        add_date_a060x=tb.add_date.isoformat() if tb.add_date else "",
                        level_a05=la,
                        level_a060x=lb,
                        a05_workbook=ta.source_workbook,
                        a05_sheet=ta.source_sheet,
                        a060x_workbook=tb.source_workbook,
                        a060x_sheet=tb.source_sheet,
                        alert=(
                            f"分级不一致：{format_traffic_source(ta)} 为 {la or '-'}，"
                            f"{format_traffic_source(tb)} 为 {lb or '-'}"
                        ),
                    )
                )

        a05_cust_dates: dict[str, list[TrafficRow]] = {}
        a060x_cust_dates: dict[str, list[TrafficRow]] = {}
        for k in a05_only_keys:
            c = a05_by_key[k].customer
            a05_cust_dates.setdefault(_norm(c), []).append(a05_by_key[k])
        for k in a060x_only_keys:
            c = a060x_by_key[k].customer
            a060x_cust_dates.setdefault(_norm(c), []).append(a060x_by_key[k])

        for key in sorted(a05_only_keys):
            ta = a05_by_key[key]
            nc = _norm(ta.customer)
            if nc in a060x_cust_dates:
                tb = a060x_cust_dates[nc][0]
                details.append(
                    TrafficCrossDetail(
                        store=store_to_label(store),
                        sales=sales,
                        customer=ta.customer,
                        status="日期不一致",
                        add_date_a05=ta.add_date.isoformat() if ta.add_date else "",
                        add_date_a060x=tb.add_date.isoformat() if tb.add_date else "",
                        level_a05=str(ta.level or ""),
                        level_a060x=str(tb.level or ""),
                        a05_workbook=ta.source_workbook,
                        a05_sheet=ta.source_sheet,
                        a060x_workbook=tb.source_workbook,
                        a060x_sheet=tb.source_sheet,
                        alert=(
                            f"同客户两边都有但添加日期不同："
                            f"{format_traffic_source(ta)}={ta.add_date}，"
                            f"{format_traffic_source(tb)}={tb.add_date}"
                        ),
                    )
                )
            else:
                peer = a060x_source_desc if a060x else "（该业务员无 A060x 新流量表）"
                details.append(
                    TrafficCrossDetail(
                        store=store_to_label(store),
                        sales=sales,
                        customer=ta.customer,
                        status="仅A05",
                        add_date_a05=ta.add_date.isoformat() if ta.add_date else "",
                        add_date_a060x="",
                        level_a05=str(ta.level or ""),
                        level_a060x="",
                        a05_workbook=ta.source_workbook,
                        a05_sheet=ta.source_sheet,
                        a060x_workbook="",
                        a060x_sheet="",
                        alert=f"仅存在于 {format_traffic_source(ta)}；对照 {peer} 无此客户",
                    )
                )

        for key in sorted(a060x_only_keys):
            tb = a060x_by_key[key]
            nc = _norm(tb.customer)
            if nc in a05_cust_dates:
                continue
            peer = a05_source_desc if a05 else "（该店铺无 A05 新流量表）"
            details.append(
                TrafficCrossDetail(
                    store=store_to_label(store),
                    sales=sales,
                    customer=tb.customer,
                    status="仅A060x",
                    add_date_a05="",
                    add_date_a060x=tb.add_date.isoformat() if tb.add_date else "",
                    level_a05="",
                    level_a060x=str(tb.level or ""),
                    a05_workbook="",
                    a05_sheet="",
                    a060x_workbook=tb.source_workbook,
                    a060x_sheet=tb.source_sheet,
                    alert=f"仅存在于 {format_traffic_source(tb)}；对照 {peer} 无此客户",
                )
            )

        a05_n, a060x_n = len(a05), len(a060x)
        if overlap_keys:
            alerts.append(f"客户+日期完全匹配 {len(overlap_keys)} 条")
        if a05_only_keys:
            only_desc = summarize_traffic_sources([a05_by_key[k] for k in a05_only_keys])
            alerts.append(f"仅 A05 侧 {len(a05_only_keys)} 条：{only_desc}")
        if a060x_only_keys:
            only_desc = summarize_traffic_sources([a060x_by_key[k] for k in a060x_only_keys])
            alerts.append(f"仅 A060x 侧 {len(a060x_only_keys)} 条：{only_desc}")
        if field_mismatch:
            alerts.append(f"重叠但分级不一致 {field_mismatch} 条（见明细）")

        if a05_n == a060x_n == len(overlap_keys) and not a05_only_keys and not a060x_only_keys and not field_mismatch:
            diff = "一致"
        elif not a05 or not a060x:
            diff = "部分缺失"
        else:
            diff = "是"

        summaries.append(
            TrafficCrossSummary(
                store=store_to_label(store),
                sales=sales,
                period=period,
                a05_source=a05_source_desc,
                a060x_source=a060x_source_desc,
                a05_count=a05_n,
                a060x_count=a060x_n,
                overlap=len(overlap_keys),
                a05_only=len(a05_only_keys),
                a060x_only=len(a060x_only_keys),
                field_mismatch=field_mismatch,
                diff=diff,
                alert="；".join(alerts) if alerts else "",
            )
        )

    return summaries, details


def _count_traffic(
    traffic_rows: list[TrafficRow],
    store: str,
    sales: str,
    month_begin: date,
    month_end: date,
    *,
    a05_only: bool = False,
) -> tuple[int, int]:
    shop = STORE_TO_SHOP.get(store, "")
    ns = _norm(sales)
    matched = [
        t
        for t in traffic_rows
        if _norm(t.sales) == ns
        and (not t.shop or t.shop == shop or shop_to_store(t.shop) == store)
        and t.add_date
        and month_begin <= t.add_date <= month_end
        and (not a05_only or is_a05_traffic(t))
    ]
    total = len(matched)
    l1plus = sum(1 for t in matched if is_l1plus(t.level))
    return total, l1plus


def _ratio(num: float | int | None, den: float | int | None) -> float | str:
    if not den:
        return ""
    return round(float(num or 0) / float(den), 4)


def _margin_pct(gross: float | None, amount: float) -> float | str:
    if gross is None or not amount:
        return ""
    return round(gross / amount, 4)


def conversion_range(cfg: dict) -> tuple[date, date, str]:
    """当月累计：月初（week.end_date 所在月 1 日）~ week.end_date。"""
    conv = cfg.get("conversion") or {}
    week_end = parse_api_date(cfg["week"]["end_date"])
    month_begin = (
        parse_api_date(conv["month_begin"])
        if conv.get("month_begin")
        else week_end.replace(day=1)
    )
    month_end = (
        parse_api_date(conv["month_end"])
        if conv.get("month_end")
        else week_end
    )
    title = conv.get("title") or (
        f"{month_begin.month}.{month_begin.day} - {month_end.month}.{month_end.day}"
    )
    return month_begin, month_end, title


def conversion_title_full(title: str) -> str:
    return f"Youro & RonChamp 业务员新客转化汇总（{title}）"


def generate_conversion_table(
    month_sales: list[dict],
    a02_match: dict[tuple, A02OrderRow],
    traffic_rows: list[TrafficRow],
    brands: dict,
    cfg: dict,
    api: YouroApi,
) -> tuple[list[list[Any]], list[OrderMetrics]]:
    conv = cfg.get("conversion") or {}
    month_begin, month_end, _ = conversion_range(cfg)
    sales_map = cfg.get("sales_name_map", {})

    metrics: list[OrderMetrics] = []
    for sale in month_sales:
        if sale.get("firstOrder") != "Y":
            continue
        od = parse_api_date(sale["orderDate"])
        if od < month_begin or od > month_end:
            continue
        sales_name = normalize_sales(sale.get("createBy", ""), sales_map)
        customer = str(sale.get("customerName", ""))
        a02 = a02_match.get((od.isoformat(), _norm(sales_name), _norm(customer)))
        store = resolve_store(a02, sale)
        traffic = find_traffic(traffic_rows, sales_name, customer, od, store)
        purchaser = api.get_purchaser_order(sale["orderNo"])
        metrics.append(compute_order_metrics(sale, purchaser, a02, traffic, brands, cfg))

    rows_out: list[list[Any]] = []
    totals = {"I": 0.0, "J": 0.0, "M": 0.0, "N": 0.0, "Q": 0.0, "R": 0.0}
    count_totals = {"D": 0, "E": 0, "G": 0, "L": 0, "P": 0}

    for store, order_list, label in (
        (STORE_YOURO, YOURO_SALES_ORDER, store_to_label(STORE_YOURO)),
        (STORE_RONCHAMP, RONCHAMP_SALES_ORDER, store_to_label(STORE_RONCHAMP)),
    ):
        sales_names: set[str] = set()
        for t in traffic_rows:
            if not is_a05_traffic(t):
                continue
            shop = t.shop or ""
            if shop_to_store(shop) == store and t.add_date and month_begin <= t.add_date <= month_end:
                sales_names.add(normalize_sales(t.sales, sales_map))
        for m in metrics:
            if m.store == store:
                sales_names.add(m.sales)
        sales_names.discard("")

        sorted_sales = sorted(sales_names, key=lambda s: _sales_sort_key(s, order_list))
        first_in_group = True
        for sales in sorted_sales:
            d, e = _count_traffic(traffic_rows, store, sales, month_begin, month_end, a05_only=True)
            tm = [m for m in metrics if m.store == store and m.sales == sales and m.channel == "tm"]
            rfq = [m for m in metrics if m.store == store and m.sales == sales and m.channel == "rfq"]
            other = [m for m in metrics if m.store == store and m.sales == sales and m.channel == "other"]
            all_deals = [m for m in metrics if m.store == store and m.sales == sales]

            tm_amt = sum(m.payment_rmb for m in tm)
            rfq_amt = sum(m.payment_rmb for m in rfq)
            other_amt = sum(m.payment_rmb for m in other)
            tm_gross = sum(m.gross or 0 for m in tm) if tm else None
            rfq_gross = sum(m.gross or 0 for m in rfq) if rfq else None
            other_gross = sum(m.gross or 0 for m in other) if other else None

            g, l_cnt, p = len(tm), len(rfq), len(other)
            c = len(all_deals)

            row = [
                label if first_in_group else "",
                sales,
                c if c else "",
                d if d else "",
                e if e else "",
                _ratio(e, d),
                g if g else "",
                _ratio(g, d),
                round(tm_amt, 2) if tm_amt else "",
                round(tm_gross, 2) if tm_gross is not None and tm else "",
                _margin_pct(tm_gross, tm_amt) if tm and tm_gross is not None else "",
                l_cnt if l_cnt else "",
                round(rfq_amt, 2) if rfq_amt else "",
                round(rfq_gross, 2) if rfq_gross is not None and rfq else "",
                _margin_pct(rfq_gross, rfq_amt) if rfq and rfq_gross is not None else "",
                p if p else "",
                round(other_amt, 2) if other_amt else "",
                round(other_gross, 2) if other_gross is not None and other else "",
                _margin_pct(other_gross, other_amt) if other and other_gross is not None else "",
                "",
            ]
            rows_out.append(row)
            first_in_group = False

            count_totals["D"] += d
            count_totals["E"] += e
            count_totals["G"] += g
            count_totals["L"] += l_cnt
            count_totals["P"] += p
            for key, col_idx in [("I", 8), ("J", 9), ("M", 12), ("N", 13), ("Q", 16), ("R", 17)]:
                val = row[col_idx]
                if val != "":
                    totals[key] += float(val)

    total_row = [
        "总计", "",
        sum(1 for m in metrics),
        count_totals["D"] or "", count_totals["E"] or "", _ratio(count_totals["E"], count_totals["D"]),
        count_totals["G"] or "", _ratio(count_totals["G"], count_totals["D"]),
        round(totals["I"], 2) if totals["I"] else "",
        round(totals["J"], 2) if totals["J"] else "",
        _margin_pct(totals["J"], totals["I"]) if totals["I"] else "",
        count_totals["L"] or "",
        round(totals["M"], 2) if totals["M"] else "",
        round(totals["N"], 2) if totals["N"] else "",
        _margin_pct(totals["N"], totals["M"]) if totals["M"] else "",
        count_totals["P"] or "",
        round(totals["Q"], 2) if totals["Q"] else "",
        round(totals["R"], 2) if totals["R"] else "",
        _margin_pct(totals["R"], totals["Q"]) if totals["Q"] else "",
        "",
    ]
    rows_out.append(total_row)
    unclassified = [m for m in metrics if m.channel == "unclassified"]
    return rows_out, unclassified


def process_week_order(
    sale: dict,
    api: YouroApi,
    a02_match: dict[tuple, A02OrderRow],
    traffic_rows: list[TrafficRow],
    brands: dict,
    cfg: dict,
) -> tuple[str, list[Any], PurchaseReview, BrandReview]:
    order_no = sale["orderNo"]
    purchaser = api.get_purchaser_order(order_no)
    od = parse_api_date(sale["orderDate"])
    sales_name = normalize_sales(sale.get("createBy", ""), cfg.get("sales_name_map", {}))
    customer = str(sale.get("customerName", ""))
    a02 = a02_match.get((od.isoformat(), _norm(sales_name), _norm(customer)))
    store = resolve_store(a02, sale)
    traffic = find_traffic(traffic_rows, sales_name, customer, od, store)
    row, pr, br = build_row(sale, purchaser, a02, traffic, brands, cfg)
    return store, row, pr, br


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate weekly new-customer order rows")
    parser.add_argument("-c", "--config", default="config.yaml")
    parser.add_argument("--write-xlsx", action="store_true", help="Write back to weekly xlsx (default: CSV only)")
    parser.add_argument("--no-conversion", action="store_true", help="Skip monthly conversion table")
    args = parser.parse_args()

    root = Path(__file__).resolve().parent
    cfg = load_config(root / args.config)
    brands = load_brands(root / cfg["paths"]["brands"])
    data_dir = root / cfg["paths"]["data_dir"]
    out_dir = root / cfg["paths"]["output_dir"]
    paths = cfg["paths"]

    api = YouroApi(cfg["api"]["base_url"], cfg["api"]["jsessionid"])
    begin = cfg["week"]["begin_date"]
    end = cfg["week"]["end_date"]
    period = cfg["week"].get("period_label") or period_label(parse_api_date(begin), parse_api_date(end))

    print(f"Fetching sales orders {begin} ~ {end} (firstOrder=Y)...")
    sales = api.list_sales_orders(begin, end)
    sales = [s for s in sales if s.get("firstOrder") == "Y"]
    print(f"  {len(sales)} first-order rows")

    print("Loading Excel sources...")
    a02_match = load_a02_orders(data_dir / paths["a02"])
    traffic_rows = load_traffic_rows(
        data_dir,
        data_dir / paths["a05"],
        paths["salesperson_globs"],
    )
    print(f"  A02 order keys: {len(a02_match)}, traffic rows: {len(traffic_rows)}")

    youro_rows: list[list[Any]] = []
    ronchamp_rows: list[list[Any]] = []
    purchase_reviews: list[PurchaseReview] = []
    brand_reviews: list[BrandReview] = []

    for sale in sorted(sales, key=lambda s: s.get("orderDate", "")):
        order_no = sale["orderNo"]
        print(f"  processing {order_no}...")
        store, row, pr, br = process_week_order(
            sale, api, a02_match, traffic_rows, brands, cfg
        )
        shop_label = STORE_TO_SHOP.get(store, store)
        print(f"    → {shop_label} ({store_to_label(store)})")
        if store == STORE_RONCHAMP:
            ronchamp_rows.append(row)
        else:
            youro_rows.append(row)
        purchase_reviews.append(pr)
        brand_reviews.append(br)

        if pr.purchase_diff in ("是", "部分缺失", "缺失"):
            print(f"    ⚠ 采购 {order_no}: {pr.purchase_diff} — {pr.purchase_alert}")

    write_csv(out_dir / "6.周新客订单表-Youro.csv", WEEKLY_HEADERS, youro_rows)
    write_csv(out_dir / "4.周新客订单表-RonChamp.csv", WEEKLY_HEADERS, ronchamp_rows)
    write_review_csv(
        out_dir / "采购核对.csv",
        ["order_no", "purchase_api", "purchase_a02_order", "purchase_diff", "purchase_alert"],
        purchase_reviews,
    )
    write_review_csv(
        out_dir / "品牌复核.csv",
        ["order_no", "brand_suggested", "brand_confidence", "brand_source", "productName_raw", "category_raw", "brand_final"],
        brand_reviews,
    )

    print(f"\nOutput:")
    print(f"  {out_dir / '6.周新客订单表-Youro.csv'} ({len(youro_rows)} rows)")
    print(f"  {out_dir / '4.周新客订单表-RonChamp.csv'} ({len(ronchamp_rows)} rows)")
    print(f"  {out_dir / '采购核对.csv'}")
    print(f"  {out_dir / '品牌复核.csv'}")

    if args.write_xlsx:
        youro_xlsx = paths.get("weekly_youro_xlsx") or paths.get("weekly_xlsx")
        if youro_xlsx:
            xlsx = data_dir / youro_xlsx
            if xlsx.exists():
                patch_weekly_xlsx(xlsx, youro_rows, period, sheet_prefix="6")
            else:
                print(f"  Youro weekly xlsx not found: {xlsx}")
        ron_xlsx = paths.get("weekly_ronchamp_xlsx")
        if ron_xlsx:
            xlsx = data_dir / ron_xlsx
            if xlsx.exists():
                patch_weekly_xlsx(xlsx, ronchamp_rows, period, sheet_prefix="4")
            else:
                print(f"  Ronchamp weekly xlsx not found: {xlsx}")

    if not args.no_conversion:
        month_begin, month_end, title = conversion_range(cfg)
        mb, me = month_begin.isoformat(), month_end.isoformat()
        print(f"\nGenerating conversion table {conversion_title_full(title)}...")
        month_sales = api.list_sales_orders(mb, me)
        conversion_rows, unclassified = generate_conversion_table(
            month_sales, a02_match, traffic_rows, brands, cfg, api
        )
        write_conversion_csv(out_dir / "2.新客转化表.csv", title, conversion_rows)
        print(f"  {out_dir / '2.新客转化表.csv'} ({len(conversion_rows)} rows, {mb} ~ {me})")
        write_review_csv(
            out_dir / "渠道未归类.csv",
            [
                "order_no", "store", "sales", "customer", "order_date",
                "payment_rmb", "gross", "traffic_file", "traffic_source", "note",
            ],
            [
                ChannelReview(
                    order_no=m.order_no,
                    store=store_to_label(m.store),
                    sales=m.sales,
                    customer=m.customer,
                    order_date=m.order_date,
                    payment_rmb=m.payment_rmb,
                    gross=m.gross,
                    traffic_file=m.traffic_file,
                    traffic_source=m.traffic_source,
                    note="无新流量匹配，未归入TM/RFQ/其他，请人工确认渠道",
                )
                for m in unclassified
            ],
        )
        if unclassified:
            print(f"  {out_dir / '渠道未归类.csv'} ({len(unclassified)} orders — 计入总成交C，不进其他P列)")

        traffic_summaries, traffic_details = compare_traffic_cross(
            traffic_rows, month_begin, month_end, cfg.get("sales_name_map", {})
        )
        write_review_csv(
            out_dir / "流量交叉核对-汇总.csv",
            [
                "store", "sales", "period", "a05_source", "a060x_source",
                "a05_count", "a060x_count", "overlap",
                "a05_only", "a060x_only", "field_mismatch", "diff", "alert",
            ],
            traffic_summaries,
        )
        write_review_csv(
            out_dir / "流量交叉核对-明细.csv",
            [
                "store", "sales", "customer", "status", "add_date_a05", "add_date_a060x",
                "level_a05", "level_a060x",
                "a05_workbook", "a05_sheet", "a060x_workbook", "a060x_sheet", "alert",
            ],
            traffic_details,
        )
        print(f"  {out_dir / '流量交叉核对-汇总.csv'} ({len(traffic_summaries)} groups)")
        print(f"  {out_dir / '流量交叉核对-明细.csv'} ({len(traffic_details)} discrepancies)")
        for s in traffic_summaries:
            if s.diff not in ("一致",):
                print(f"    ⚠ 流量 {s.store}/{s.sales}: {s.diff} — {s.alert}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
